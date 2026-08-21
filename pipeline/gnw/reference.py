"""Reference datasets the flag engine cross-references against.

  nppes       monthly full replacement (active NPIs) -> nppes.parquet
              + Monthly Deactivation Report -> nppes_deactivated.parquet
              (deactivated NPIs are REMOVED from the main file; the xlsx
              report is the only public deactivation source, and the free
              API cannot see deactivated NPIs at all)
  pufs        CMS Plan Attributes + Service Area PUFs -> plan_attributes.parquet,
              service_area.parquet
  landscape   QHP landscape (individual medical) -> plan_county.parquet
              (one row per plan per county; the MVP county denominator)
  nucc        NUCC taxonomy (archived at scoping time) -> nucc_taxonomy.parquet
  zcta        Census ZCTA<->county relationship -> zip_county.parquet
              (HUD's ZIP-COUNTY crosswalk is better but needs an account
              token; Census is free and adequate for MVP)

Every download is recorded in data/reference/sources.json with sha256,
size, and Last-Modified — reference data backs published claims, so it
gets provenance like everything else.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
import requests

from .quirks import BROWSER_UA

log = logging.getLogger("gnw.reference")

NPPES_LISTING = "https://download.cms.gov/nppes/NPI_Files.html"
PUF_PLAN_ATTRIBUTES = "https://download.cms.gov/marketplace-puf/2026/plan-attributes-puf.zip"
PUF_SERVICE_AREA = "https://download.cms.gov/marketplace-puf/2026/service-area-puf.zip"
LANDSCAPE_MEDICAL = "https://data.healthcare.gov/datafile/py2026/individual_market_medical.zip"
CENSUS_ZCTA_COUNTY = (
    "https://www2.census.gov/geo/docs/maps-data/data/rel2020/zcta520/"
    "tab20_zcta520_county20_natl.txt"
)
NUCC_ARCHIVED = Path(__file__).resolve().parents[2] / "scoping" / "evidence" / "nucc_taxonomy_261.csv"


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


class ReferenceStore:
    def __init__(self, root: Path) -> None:
        self.raw = root / "raw"
        self.parquet = root / "parquet"
        self.sources_path = root / "sources.json"
        self.raw.mkdir(parents=True, exist_ok=True)
        self.parquet.mkdir(parents=True, exist_ok=True)

    def record(self, name: str, url: str, path: Path, last_modified: str | None) -> None:
        sources = {}
        if self.sources_path.exists():
            sources = json.loads(self.sources_path.read_text())
        h = hashlib.sha256()
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1 << 20), b""):
                h.update(chunk)
        sources[name] = {
            "url": url,
            "file": str(path),
            "sha256": h.hexdigest(),
            "bytes": path.stat().st_size,
            "last_modified": last_modified,
            "fetched_at": _now(),
        }
        self.sources_path.write_text(json.dumps(sources, indent=1))

    def download(self, name: str, url: str, filename: str | None = None) -> Path:
        dest = self.raw / (filename or url.rsplit("/", 1)[-1])
        resp = requests.get(
            url, headers={"User-Agent": BROWSER_UA}, stream=True, timeout=(15, 600)
        )
        resp.raise_for_status()
        with open(dest, "wb") as fh:
            for chunk in resp.iter_content(1 << 20):
                fh.write(chunk)
        self.record(name, url, dest, resp.headers.get("Last-Modified"))
        log.info("downloaded %s: %s (%.1f MB)", name, dest.name, dest.stat().st_size / 1e6)
        return dest


def _write_parquet(path: Path, rows: list[dict], schema: pa.Schema) -> None:
    pq.write_table(pa.Table.from_pylist(rows, schema=schema), path, compression="zstd")
    log.info("wrote %s: %d rows", path.name, len(rows))


# --- NPPES -------------------------------------------------------------------


def nppes_current_urls() -> tuple[str, str]:
    """Scrape the listing page for the current monthly + deactivation zips."""
    resp = requests.get(NPPES_LISTING, headers={"User-Agent": BROWSER_UA}, timeout=60)
    resp.raise_for_status()
    base = NPPES_LISTING.rsplit("/", 1)[0]

    def find(pattern: str) -> str:
        m = re.search(pattern, resp.text)
        if not m:
            raise RuntimeError(f"pattern {pattern!r} not on NPPES listing page")
        href = m.group(0)
        return href if href.startswith("http") else f"{base}/{href.lstrip('./')}"

    monthly = find(r"[\w./]*NPPES_Data_Dissemination_\w+_\d{4}_V2\.zip")
    deact = find(r"[\w./]*NPPES_Deactivated_NPI_Report_\d{6}_V2\.zip")
    return monthly, deact


_NPPES_KEEP = [
    "NPI",
    "Entity Type Code",
    "Provider Organization Name (Legal Business Name)",
    "Provider Last Name (Legal Name)",
    "Provider First Name",
    "Provider Business Practice Location Address City Name",
    "Provider Business Practice Location Address State Name",
    "Provider Business Practice Location Address Postal Code",
    "Provider Business Practice Location Address Telephone Number",
    "NPI Deactivation Date",
    "NPI Reactivation Date",
    "Last Update Date",
]
_NPPES_TAXONOMY = [f"Healthcare Provider Taxonomy Code_{i}" for i in range(1, 16)]
_NPPES_PRIMARY = [f"Healthcare Provider Primary Taxonomy Switch_{i}" for i in range(1, 16)]

NPPES_SCHEMA = pa.schema(
    [
        ("npi", pa.string()),
        ("entity_type", pa.string()),
        ("org_name", pa.string()),
        ("last_name", pa.string()),
        ("first_name", pa.string()),
        ("practice_city", pa.string()),
        ("practice_state", pa.string()),
        ("practice_zip", pa.string()),
        ("practice_phone", pa.string()),
        ("deactivation_date", pa.string()),
        ("reactivation_date", pa.string()),
        ("last_update", pa.string()),
        ("taxonomy_primary", pa.string()),
        ("taxonomies", pa.string()),
    ]
)


def build_nppes(store: ReferenceStore) -> None:
    monthly_url, deact_url = nppes_current_urls()

    # Deactivation report: the sole public deactivation source.
    deact_zip = store.download("nppes_deactivated", deact_url)
    import openpyxl

    with zipfile.ZipFile(deact_zip) as zf:
        xlsx_name = next(n for n in zf.namelist() if n.endswith(".xlsx"))
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(xlsx_name)), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows, header_seen = [], False
    for row in ws.iter_rows(values_only=True):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if not header_seen:
            if vals and vals[0].upper() == "NPI":
                header_seen = True
            continue
        if vals and re.fullmatch(r"\d{10}", vals[0]):
            rows.append({"npi": vals[0], "deactivation_date": vals[1] if len(vals) > 1 else None})
    _write_parquet(
        store.parquet / "nppes_deactivated.parquet",
        rows,
        pa.schema([("npi", pa.string()), ("deactivation_date", pa.string())]),
    )

    # Monthly full replacement: stream the 11.6GB CSV out of the zip.
    monthly_zip = store.download("nppes_monthly", monthly_url)
    out = store.parquet / "nppes.parquet"
    writer = pq.ParquetWriter(out, NPPES_SCHEMA, compression="zstd")
    batch: list[dict] = []
    n = 0
    with zipfile.ZipFile(monthly_zip) as zf:
        main = next(
            n_ for n_ in zf.namelist() if n_.startswith("npidata_pfile") and "fileheader" not in n_
        )
        with zf.open(main) as raw:
            text = io.TextIOWrapper(raw, encoding="utf-8", errors="replace", newline="")
            reader = csv.reader(text)
            header = next(reader)
            idx = {}
            for col in _NPPES_KEEP + _NPPES_TAXONOMY + _NPPES_PRIMARY:
                if col not in header:
                    raise RuntimeError(f"NPPES column {col!r} missing; layout changed?")
                idx[col] = header.index(col)
            tax_idx = [(idx[c], idx[p]) for c, p in zip(_NPPES_TAXONOMY, _NPPES_PRIMARY)]
            for rec in reader:
                taxonomies, primary = [], None
                for c_i, p_i in tax_idx:
                    code = rec[c_i]
                    if code:
                        taxonomies.append(code)
                        if rec[p_i] == "Y" and primary is None:
                            primary = code
                batch.append(
                    {
                        "npi": rec[idx["NPI"]],
                        "entity_type": rec[idx["Entity Type Code"]] or None,
                        "org_name": rec[idx["Provider Organization Name (Legal Business Name)"]] or None,
                        "last_name": rec[idx["Provider Last Name (Legal Name)"]] or None,
                        "first_name": rec[idx["Provider First Name"]] or None,
                        "practice_city": rec[idx["Provider Business Practice Location Address City Name"]] or None,
                        "practice_state": rec[idx["Provider Business Practice Location Address State Name"]] or None,
                        "practice_zip": rec[idx["Provider Business Practice Location Address Postal Code"]] or None,
                        "practice_phone": rec[idx["Provider Business Practice Location Address Telephone Number"]] or None,
                        "deactivation_date": rec[idx["NPI Deactivation Date"]] or None,
                        "reactivation_date": rec[idx["NPI Reactivation Date"]] or None,
                        "last_update": rec[idx["Last Update Date"]] or None,
                        "taxonomy_primary": primary or (taxonomies[0] if taxonomies else None),
                        "taxonomies": "|".join(taxonomies) or None,
                    }
                )
                n += 1
                if len(batch) >= 200_000:
                    writer.write_table(pa.Table.from_pylist(batch, schema=NPPES_SCHEMA))
                    batch = []
    if batch:
        writer.write_table(pa.Table.from_pylist(batch, schema=NPPES_SCHEMA))
    writer.close()
    log.info("wrote nppes.parquet: %d active NPIs", n)


# --- CMS PUFs + landscape ----------------------------------------------------


def _csv_from_zip(zip_path: Path) -> tuple[list[str], csv.reader]:
    zf = zipfile.ZipFile(zip_path)
    name = next(n for n in zf.namelist() if n.lower().endswith(".csv"))
    text = io.TextIOWrapper(zf.open(name), encoding="utf-8-sig", errors="replace", newline="")
    reader = csv.reader(text)
    return next(reader), reader


def build_pufs(store: ReferenceStore) -> None:
    pa_zip = store.download("plan_attributes_puf", PUF_PLAN_ATTRIBUTES)
    header, reader = _csv_from_zip(pa_zip)
    want = [
        "BusinessYear", "StateCode", "IssuerId", "StandardComponentId", "PlanId",
        "PlanMarketingName", "MetalLevel", "DentalOnlyPlan", "ServiceAreaId",
        "MarketCoverage", "PlanType",
    ]
    idx = {c: header.index(c) for c in want if c in header}
    missing = [c for c in want if c not in idx]
    if missing:
        raise RuntimeError(f"Plan Attributes PUF missing columns {missing}; header={header[:20]}")
    rows = [
        {k.lower(): (rec[i] or None) for k, i in idx.items()}
        for rec in reader
        if len(rec) >= len(header)
    ]
    schema = pa.schema([(k.lower(), pa.string()) for k in want])
    _write_parquet(store.parquet / "plan_attributes.parquet", rows, schema)

    sa_zip = store.download("service_area_puf", PUF_SERVICE_AREA)
    header, reader = _csv_from_zip(sa_zip)
    want = [
        "BusinessYear", "StateCode", "IssuerId", "ServiceAreaId", "ServiceAreaName",
        "CoverEntireState", "County", "PartialCounty", "ZipCodes", "MarketCoverage",
    ]
    idx = {c: header.index(c) for c in want if c in header}
    missing = [c for c in want if c not in idx]
    if missing:
        raise RuntimeError(f"Service Area PUF missing columns {missing}; header={header[:20]}")
    rows = [
        {k.lower(): (rec[i] or None) for k, i in idx.items()}
        for rec in reader
        if len(rec) >= len(header)
    ]
    schema = pa.schema([(k.lower(), pa.string()) for k in want])
    _write_parquet(store.parquet / "service_area.parquet", rows, schema)


def build_landscape(store: ReferenceStore) -> None:
    """QHP landscape: one row per plan per county — the county denominator."""
    zip_path = store.download("landscape_medical", LANDSCAPE_MEDICAL)
    import openpyxl

    with zipfile.ZipFile(zip_path) as zf:
        xlsx_name = next(n for n in zf.namelist() if n.endswith(".xlsx"))
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(xlsx_name)), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = None
    for row in rows_iter:  # header is on row 2; scan for the row containing 'Plan ID'
        vals = [str(v).strip() if v is not None else "" for v in row]
        if any("Plan ID" in v for v in vals):
            header = vals
            break
    if header is None:
        raise RuntimeError("landscape header row not found")

    def col(fragment: str) -> int:
        for i, name in enumerate(header):
            if fragment.lower() in name.lower():
                return i
        raise RuntimeError(f"landscape column ~{fragment!r} not found; header={header[:15]}")

    c_state, c_county, c_fips = col("State"), col("County Name"), col("FIPS")
    c_issuer, c_plan = col("Issuer Name"), col("Plan ID")
    c_metal, c_name, c_type = col("Metal Level"), col("Plan Marketing Name"), col("Plan Type")
    rows = []
    for row in rows_iter:
        vals = [str(v).strip() if v is not None else None for v in row]
        if not vals or not vals[c_plan] or len(str(vals[c_plan])) < 14:
            continue
        rows.append(
            {
                "state": vals[c_state],
                "county_name": vals[c_county],
                "fips": vals[c_fips],
                "issuer_name": vals[c_issuer],
                "plan_id": vals[c_plan],
                "metal_level": vals[c_metal],
                "plan_marketing_name": vals[c_name],
                "plan_type": vals[c_type],
            }
        )
    schema = pa.schema(
        [(f, pa.string()) for f in
         ("state", "county_name", "fips", "issuer_name", "plan_id", "metal_level",
          "plan_marketing_name", "plan_type")]
    )
    _write_parquet(store.parquet / "plan_county.parquet", rows, schema)


# --- NUCC + ZIP->county ------------------------------------------------------


def build_nucc(store: ReferenceStore) -> None:
    with open(NUCC_ARCHIVED, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        rows = [
            {
                "code": r.get("Code"),
                "grouping": r.get("Grouping"),
                "classification": r.get("Classification"),
                "specialization": r.get("Specialization"),
            }
            for r in reader
            if r.get("Code")
        ]
    schema = pa.schema(
        [("code", pa.string()), ("grouping", pa.string()),
         ("classification", pa.string()), ("specialization", pa.string())]
    )
    _write_parquet(store.parquet / "nucc_taxonomy.parquet", rows, schema)
    store.record("nucc_taxonomy", "archived: scoping/evidence/nucc_taxonomy_261.csv", NUCC_ARCHIVED, None)


def build_zcta(store: ReferenceStore) -> None:
    path = store.download("census_zcta_county", CENSUS_ZCTA_COUNTY)
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter="|")
        for r in reader:
            z, fips = r.get("GEOID_ZCTA5_20"), r.get("GEOID_COUNTY_20")
            if z and fips:
                rows.append({"zip": z, "county_fips": fips})
    _write_parquet(
        store.parquet / "zip_county.parquet",
        rows,
        pa.schema([("zip", pa.string()), ("county_fips", pa.string())]),
    )


BUILDERS = {
    "nppes": build_nppes,
    "pufs": build_pufs,
    "landscape": build_landscape,
    "nucc": build_nucc,
    "zcta": build_zcta,
}
