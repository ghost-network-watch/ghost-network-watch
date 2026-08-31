"""Turn callsheet.csv into an easy-to-use workbook for the actual calls.

The full CSV carries analysis columns (NPI, plan ID, cell grade, ...) that
get in the way while dialing. This produces callsheet_easy.xlsx with three
tabs: instructions, a dial-friendly Calls tab with one outcome dropdown, and
a Reference tab holding every original column so nothing is lost.

The Calls tab deliberately hides the cell grade: a caller who knows a row
came from an F cell hears what they expect to hear. Grades stay in the
Reference tab for analysis after the calls are done.

Usage: python ops/make_easy_callsheet.py [snapshot]  (default 2026-08)
Output sits next to the input CSV. Private research instrument; never
publish or commit it (data/ is gitignored).
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(__file__).resolve().parents[1]

OUTCOMES = [
    "no number published",
    "dead line / wrong number",
    "no answer / voicemail",
    "provider unknown here",
    "not in network",
    "not taking new patients",
    "taking new patients",
    "unclear (see notes)",
]

INSTRUCTIONS = [
    ("Ghost Network Watch phone study, the short version", True),
    ("", False),
    ("One row is one call, and no phone number appears twice.", False),
    ("", False),
    ("Say: \"I'm shopping for a marketplace plan and I'd like to see", False),
    ("[Ask for]. Which insurers are they in network with, and are they", False),
    ("taking new patients?\"", False),
    ("", False),
    ("Let the office name its own carriers instead of reading plan names at", False),
    ("them. Put what they say in 'Carriers they named', then check whether", False),
    ("'Carrier to check' is in that list. Plan variants inside one carrier", False),
    ("share a network, so the carrier-level answer is the answer.", False),
    ("", False),
    ("Then fill three cells: Called on, Outcome (pick from the dropdown,", False),
    ("the furthest you got down this ladder), and Notes if anything was odd.", False),
    ("", False),
    ("The outcome ladder:", True),
    ("  no number published        pre-filled; there is nothing to dial", False),
    ("  dead line / wrong number   it never rang, or a fax, or not an office", False),
    ("  no answer / voicemail      rang but no human; try once more another day", False),
    ("  provider unknown here      a human answered, never heard of them", False),
    ("  not in network             provider is real but not in that plan", False),
    ("  not taking new patients    in network, panel closed", False),
    ("  taking new patients        ask how many days until the first opening,", False),
    ("                             put the number in 'Appt in days', don't book", False),
    ("  unclear (see notes)        anything that fits nowhere; explain in Notes", False),
    ("", False),
    ("Ground rules: business hours in the row's time zone, let it ring a", False),
    ("full minute, stay honest, keep it short, never hold an appointment.", False),
]


def fmt_phone(p: str) -> str:
    digits = "".join(c for c in p if c.isdigit())
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return p


def undialable(p: str) -> bool:
    digits = "".join(c for c in p if c.isdigit())
    if len(digits) != 10:
        return True
    return len(set(digits)) == 1 or digits[:3] in ("999", "000", "555")


def main() -> None:
    snapshot = sys.argv[1] if len(sys.argv) > 1 else "2026-08"
    src = REPO / "data" / "callsheet" / snapshot / "callsheet.csv"
    rows = [r for r in csv.DictReader(src.open()) if r["study_id"] != "EXAMPLE"]
    # One office often serves several rows; group them so it's one call.
    rows.sort(key=lambda r: (r["addr_state"], r["phone"], r["provider_name"]))

    wb = Workbook()

    ws = wb.active
    ws.title = "How to call"
    ws.column_dimensions["A"].width = 78
    for text, bold in INSTRUCTIONS:
        ws.append([text])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    calls = wb.create_sheet("Calls")
    # "Carrier to check" is the insurer, not the plan variant: plan variants
    # inside one carrier share a network, and a front desk answers at the
    # carrier level anyway.
    header = ["#", "Ask for", "Carrier to check", "City, ST", "Phone",
              "Called on", "Outcome", "Carriers they named", "Appt in days", "Notes"]
    widths = [5, 26, 30, 20, 16, 11, 24, 26, 12, 40]
    calls.append(header)
    for i, w in enumerate(widths, 1):
        calls.column_dimensions[get_column_letter(i)].width = w
        c = calls.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    calls.freeze_panes = "A2"

    dv = DataValidation(
        type="list", formula1='"' + ",".join(OUTCOMES) + '"', allow_blank=True
    )
    calls.add_data_validation(dv)

    fill_todo = PatternFill("solid", fgColor="FFF6DD")
    for r in rows:
        dead = undialable(r["phone"])
        calls.append([
            int(r["study_id"]),
            r["provider_name"].title(),
            r["issuer_name"],
            f'{r["city"].title()}, {r["addr_state"]}',
            fmt_phone(r["phone"]),
            None,
            OUTCOMES[0] if dead else None,
            None,
            None,
            "nothing to dial" if dead else None,
        ])
        n = calls.max_row
        dv.add(calls.cell(row=n, column=7))
        if not dead:
            for col in (6, 7):
                calls.cell(row=n, column=col).fill = fill_todo
        calls.cell(row=n, column=10).alignment = Alignment(wrap_text=True)

    ref = wb.create_sheet("Reference (analysis only)")
    with src.open() as f:
        for row in csv.reader(f):
            ref.append(row)
    ref.freeze_panes = "A2"
    for c in ref[1]:
        c.font = Font(bold=True)

    out = src.with_name("callsheet_easy.xlsx")
    wb.save(out)
    print(f"{out}  ({len(rows)} rows, "
          f"{sum(1 for r in rows if undialable(r['phone']))} pre-filled as undialable)")


if __name__ == "__main__":
    main()
